
#Bibliotecas
import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pyautogui import alert as al
import os.path
from openpyxl import load_workbook as lw

def calculo():
    caminho_base = str(input("Caminho do arquivo - Origem:"))
    caminho_saida = str(input("Caminho do arquivo - Saida:"))
    al('Iniciando apuração!')
    try:
        dados = os.listdir(rf"{caminho_base}")
        for arquivos in dados:
            if '.xlsx' in arquivos:
                try:
                    tabela = pd.read_excel(rf"{caminho_base}\{arquivos}")
                    tabela['Comissao'] = tabela['Valor Venda']*(20/100)
                    diretorio = rf"{caminho_saida}"
                    file = "Nova apuração - " + arquivos
                    tabela.to_excel(os.path.join(diretorio, file), index=False)
                except:
                    al(f'Deu erro no arquivo: {arquivos}')
                dados_2 = os.listdir(rf"{caminho_saida}")
                for arquivo in dados_2:
                    if '.xlsx' in arquivo:
                        try:
                            wb = load_workbook(rf'{caminho_saida}\{arquivo}')
                            sheet = wb['Sheet1']
                            # coluna que começa
                            min_column = wb.active.min_column
                            # coluna que termina
                            max_column = wb.active.max_column
                            # linha que começa
                            min_row = wb.active.min_row
                            # linha que termina
                            max_row = wb.active.max_row
                            sheet[f'F2'] = f" Total "
                            sheet[f'G2'] = f"=SUM(d2:d{max_row})"
                            cell = sheet[f'F2']
                            cell.number_format = 'R$ #,##0.00'  # Formato para Real Brasileiro
                            sheet.column_dimensions[f'F'].width = 20
                            wb.save(rf'{caminho_saida}\{arquivo}')
                            wb.close()
                        except:
                            al(f'Deu erro no arquivo: {arquivo}')
        al('Apuração Finalizada!')
    except:
        al('Caminho não especificado')


calculo()